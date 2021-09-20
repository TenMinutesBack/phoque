from openpyxl import load_workbook, Workbook
import xml.etree.ElementTree as ET
import os

def action_examiner(case_dir):
    root = ET.parse(case_dir).getroot()
    if 'action' in [elem.tag for elem in root.iter()]:
        return True

def no_precode(case_dir):
    if not ET.parse(case_dir).getroot().attrib['preConditions']:
        return True

def determine_week(wb, sheet, week):
    for title in wb[sheet].iter_rows(max_row=1, values_only=True):
        for count, value in enumerate(title):
            if value == week:
                return count + 1

def path_finder(tcid):
    root_10 = r'C:\Users\jeter_lin\Documents\Gerrit\src\1080p_tt\scripts\bj'
    root_13 = r'C:\Users\jeter_lin\Documents\Gerrit\src\2400x960_tt\scripts\bj'
    for root, directory, filename in os.walk(root_10):
        if tcid + '.xml' in filename:
            return root + '\\' + tcid + '.xml'
    for root, directory, filename in os.walk(root_13):
        if tcid + '.xml' in filename:
            return root + '\\' + tcid + '.xml'

if __name__ == '__main__':
    new_wb = Workbook()
    ws = new_wb.active

    headers = ['tcid', 'pass rate', 'no precode', 'has action']

    contains_action_and_fail = new_wb.create_sheet('contains_action_and_fail')
    contains_action_but_pass = new_wb.create_sheet('contains_action_but_pass')
    # For case that has pass rate less than 80%
    needs_special_care = new_wb.create_sheet('needs_special_care')

    wb = load_workbook('one_month.xlsx',data_only=True)
    
    for sheet in wb.sheetnames:
        rlt_index = determine_week(wb, sheet, 'W37')
        if rlt_index:
            for case in wb[sheet].iter_rows(max_col=rlt_index, values_only=True):
                if case[0] and str(case[0]).lower()[:3] == 'tc_':
                    path = path_finder(case[0].lower())
                    if path:
                        # determine the action
                        contains_action = action_examiner(path)
                        # determine the precode
                        if no_precode(path):
                            did_not_has_precode = True
                        else:
                            did_not_has_precode =''
                        
                        case_detail = [case[0], round(case[1], 2), did_not_has_precode, contains_action]

                        ws.append(case_detail)

                                
    new_wb.save('script_contains_action.xlsx')

    # for i in wb['waa'].iter_rows(max_col=2, max_row=2, values_only=True):
    #     print(type(i[1]))