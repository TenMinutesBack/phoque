from openpyxl import load_workbook, Workbook
import xml.etree.ElementTree as ET
import os

def path_finder(tcid):
    root_10 = r'C:\Users\jeter_lin\Documents\Gerrit\src\1080p_tt\scripts\bj'
    root_13 = r'C:\Users\jeter_lin\Documents\Gerrit\src\2400x960_tt\scripts\bj'
    for root, directory, filename in os.walk(root_10):
        if tcid + '.xml' in filename:
            return root + '\\' + tcid + '.xml'
    for root, directory, filename in os.walk(root_13):
        if tcid + '.xml' in filename:
            return root + '\\' + tcid + '.xml'

def get_precode(tc_dir):
    return ET.parse(tc_dir).getroot().attrib['preConditions'].split(',')

def action_examiner(case_dir):
    root = ET.parse(case_dir).getroot()
    if 'action' in [elem.tag for elem in root.iter()]:
        return 'V'

if __name__ == '__main__':
    pkg_design = load_workbook('new_one_month.xlsx')
    wb = Workbook()
    summary = wb.create_sheet('Summary')

    title = ['TCID', 'Root', 'Action?', 'Precode']
    
    for name in pkg_design.sheetnames:
        print('iterating {}'.format(name))
        if name not in ['Summary', 'back_to_manual', 'Taipei AI Scope']:
            wb.create_sheet(name)
            wb[name].append(title)
            pkg_precode_summary = []
            for tcid in pkg_design[name].iter_rows(max_col=1, values_only=True):
                if tcid[0] and tcid[0][:3].lower() == 'tc_': 
                    path = path_finder(tcid[0].lower())
                    if path:
                        action = action_examiner(path)
                        precode = get_precode(path)
                    
                    tmp = [tcid[0], path, action] + sorted(precode)

                    for code in precode:
                        if code not in pkg_precode_summary:
                            pkg_precode_summary.append(code)

                    wb[name].append(tmp)
                
            wb['Summary'].append([name] + pkg_precode_summary)
        
    wb.save('case_detail.xlsx')